""" Download SAP table and insert into spreadsheet """
import argparse
import getpass
from datetime import date
from datetime import timedelta
from pyrfc import Connection
from openpyxl import load_workbook

today = date.today()
yesterday = today - timedelta(days = 1)
from_ts = yesterday.strftime("%Y%m%d%H%M%S")
to_ts = today.strftime("%Y%m%d") + '000000'

parser = argparse.ArgumentParser(description='Download SAP table and insert into spreadsheet',
    formatter_class=argparse.ArgumentDefaultsHelpFormatter)
parser.add_argument('-f', '--file', help='Excel spreadsheet file',
    default='mypivot.xlsx')
parser.add_argument('-d', '--data', help='Excel spreadsheet data tab', default='data')
parser.add_argument('-u', '--user', help='SAP user', default='MYUSER')
parser.add_argument('-r', '--rfc', help='SAP RFC function module', default='/SAPDS/RFC_READ_TABLE')
parser.add_argument('-o', '--options',
    help='SAP RFC function module options',
    default='CONFIRMED_AT_WH GE ' + from_ts + ' AND CONFIRMED_AT_WH LT ' +
        to_ts + " AND LGNUM EQ 'WH01'")
parser.add_argument('-i', '--fields', help='SAP table fields', nargs='*',
    default=['TANUM', 'TAPOS', 'PROCTY', 'CONFIRMED_AT_WH'])
parser.add_argument('-j', '--confirmation', help='Task confirmation time',
    default='Confirmation Time in Warehouse Time Zone')
parser.add_argument('-g', '--shift', help='Morning shift end time',
    default=151500)
parser.add_argument('-m', '--morning', help='Morning shift column',
    default='Morning shift')
parser.add_argument('-t', '--table', help='SAP table', default='/SCWM/ORDIM_C')
parser.add_argument('-a', '--ashost', help='SAP host', default='MYSAPHOST')
parser.add_argument('-n', '--sysnr', help='SAP system number', default='00')
parser.add_argument('-s', '--sysid', help='SAP system ID', default='PRD')
parser.add_argument('-c', '--client', help='SAP client', default='100')
parser.add_argument('-l', '--language', help='SAP login language', default='EN')
parser.add_argument('-e', '--delimiter', help='Internal column delimiter', default='|')
parser.add_argument('-p', '--columns', help='Data tab columns to delete', default=200)

args = parser.parse_args()
passwd = getpass.getpass(prompt='SAP password for ' + args.user + ':')

abap_system = {
    'user'      : args.user,
    'passwd'    : passwd,
    'ashost'    : args.ashost,
    'sysnr'     : args.sysnr,
    'sysid'     : args.sysid,
    'client'    : args.client,
    'lang'      : args.language
    }

conn = Connection(**abap_system)
ordim_c = conn.call(args.rfc, QUERY_TABLE = args.table,
    DELIMITER = args.delimiter, FIELDS = args.fields,
    OPTIONS = [args.options])

wb = load_workbook(filename = args.file)
if args.data in wb:
    wbsd = wb[args.data]
    wbsd.delete_cols(1, args.columns)
else:
    wbsd = wb.create_sheet(title=args.data)

ordim_col_labs = [x['FIELDTEXT'] for x in ordim_c['FIELDS']]
for idx, val in enumerate(ordim_col_labs):
    wbsd.cell(column=idx+1, row=1, value=val)
    if val == args.confirmation:
        conf_col = idx
        ordim_col_labs.append(args.morning)
ordim_data = [x['WA'].split(args.delimiter) for x in ordim_c['DATA']]
for idx, val in enumerate(ordim_data):
    for idx2, val2 in enumerate(val):
        if idx2 == conf_col and int(val2[9:]) < args.shift:
            val.append('X')
        wbsd.cell(column=idx2+1, row=idx+2, value=val2.strip())

wb.save(filename = args.file)
print(args)
